import webbrowser
from jinja2 import debug
from google.ai.generativelanguage_v1beta.services.cache_service import pagers
from pypdf import pagerange
import os
import datetime
import openpyxl
import re
import json
import asyncio
import base64
from io import BytesIO
import pandas as pd
import google.generativeai as genai
from copy import copy
from jinja2 import Environment, BaseLoader, StrictUndefined, Undefined
from playwright.async_api import async_playwright
from pypdf import PdfReader, PdfWriter
from components.config import AppConfig

class ExcelService:
    def __init__(self, config: AppConfig):
        self.config = config

    def save_submission(self, full_name: str, phone: str, email: str, 
                        property_type: str, suburb_input: str, 
                        budget: str, intention: str, priorities_yes_no: list) -> str:
        excel_path = self.config.excel_path
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Find the last row with data in Column A (Submission ID)
        last_row = 3
        for r in range(4, sheet.max_row + 2):
            if sheet.cell(row=r, column=1).value is not None:
                last_row = r
                
        # Generate Submission ID
        last_id = sheet.cell(row=last_row, column=1).value
        if last_id and isinstance(last_id, str) and last_id.startswith("SPG-"):
            try:
                num = int(last_id.split("-")[1])
                next_id = f"SPG-{num + 1:03d}"
            except (IndexError, ValueError):
                next_id = "SPG-002"
        else:
            next_id = "SPG-001"
            
        # Parse Suburb, Postcode, and State
        suburb_clean, postcode_clean, state_clean = "", "", ""
        if suburb_input:
            postcode_match = re.search(r"\b\d{3,4}\b", suburb_input)
            postcode_clean = postcode_match.group(0) if postcode_match else ""
            
            state_match = re.search(r"\b(VIC|NSW|QLD|WA|SA|TAS|ACT|NT)\b", suburb_input, re.IGNORECASE)
            state_clean = state_match.group(0).upper() if state_match else ""
            
            suburb_clean = suburb_input
            if postcode_clean:
                suburb_clean = suburb_clean.replace(postcode_clean, "")
            if state_clean:
                suburb_clean = re.sub(rf"\b{state_clean}\b", "", suburb_clean, flags=re.IGNORECASE)
                
            suburb_clean = re.sub(r"[,\-\s]+", " ", suburb_clean).strip()
            
        next_row = last_row + 1
        date_submitted = datetime.date.today().strftime("%d/%m/%Y")
        
        # Assemble row data (25 columns)
        row_values = [
            next_id,            # 1: Submission ID
            date_submitted,     # 2: Date Submitted
            full_name,          # 3: Full Name
            email,              # 4: Email Address
            phone,              # 5: Phone Number
            property_type,      # 6: Property Type
            suburb_clean,       # 7: Suburb / Area
            postcode_clean,     # 8: Postcode
            state_clean,        # 9: State
            budget,             # 10: Budget Range
            intention,          # 11: Buying Purpose
        ]
        # Append 11 priorities
        row_values.extend(priorities_yes_no)
        # Append remaining columns: Additional Notes (23), Report Status (24), Assigned To (25)
        row_values.extend(["", "Pending", "Shoyeb"])
        
        # Write to the cells and copy style if last_row has styles
        for col_idx, val in enumerate(row_values, start=1):
            new_cell = sheet.cell(row=next_row, column=col_idx, value=val)
            if last_row >= 4:
                src_cell = sheet.cell(row=last_row, column=col_idx)
                if src_cell.has_style:
                    new_cell.font = copy(src_cell.font)
                    new_cell.border = copy(src_cell.border)
                    new_cell.fill = copy(src_cell.fill)
                    new_cell.number_format = copy(src_cell.number_format)
                    new_cell.protection = copy(src_cell.protection)
                    new_cell.alignment = copy(src_cell.alignment)
                    
        wb.save(excel_path)
        return next_id


class DataService:
    def __init__(self, config: AppConfig):
        self.config = config

    def filter_property_data(self, df: pd.DataFrame, postcode_str: str, budget_str: str, property_type_str: str) -> pd.DataFrame:
        if df is None or len(df) == 0:
            return df
        df_filtered = df.copy()
        if postcode_str:
            try:
                pc_val = float(postcode_str)
                if 'Property post code' in df_filtered.columns:
                    df_filtered = df_filtered[df_filtered['Property post code'] == pc_val]
            except ValueError:
                pass
        if budget_str and 'Purchase price' in df_filtered.columns:
            if "Under $500k" in budget_str:
                df_filtered = df_filtered[df_filtered['Purchase price'] < 500000]
            elif "$500k" in budget_str and "$800k" in budget_str:
                df_filtered = df_filtered[(df_filtered['Purchase price'] >= 500000) & (df_filtered['Purchase price'] <= 800000)]
            elif "$800k" in budget_str and "$1.2M" in budget_str:
                df_filtered = df_filtered[(df_filtered['Purchase price'] >= 800000) & (df_filtered['Purchase price'] <= 1200000)]
            elif "Above $1.2M" in budget_str:
                df_filtered = df_filtered[df_filtered['Purchase price'] > 1200000]
        if property_type_str and 'Primary purpose' in df_filtered.columns:
            if property_type_str == "Land":
                df_filtered = df_filtered[df_filtered['Primary purpose'] == 'Vacant land']
            elif property_type_str in ["House", "Unit", "Townhouse"]:
                df_filtered = df_filtered[df_filtered['Primary purpose'] == 'Residence']
        if len(df_filtered) > 50:
            if 'Contract date' in df_filtered.columns:
                try:
                    df_filtered = df_filtered.sort_values(by='Contract date', ascending=False)
                except Exception:
                    pass
            df_filtered = df_filtered.head(50)
        return df_filtered

    def auto_load_postcode_dataset(self, postcode_str: str):
        if not postcode_str:
            return None, None
        try:
            pc_val = float(postcode_str)
            split_dir = self.config.split_dir
            if os.path.exists(split_dir):
                for fname in os.listdir(split_dir):
                    if fname.startswith("postcode_") and fname.endswith(".csv"):
                        parts = fname.replace("postcode_", "").replace(".csv", "").split("_to_")
                        if len(parts) == 2:
                            start_pc = float(parts[0])
                            end_pc = float(parts[1])
                            if start_pc <= pc_val <= end_pc:
                                csv_path = os.path.join(split_dir, fname)
                                return pd.read_csv(csv_path), fname
        except Exception as e:
            raise RuntimeError(f"Could not auto-load postcode dataset: {e}")
        return None, None


class GeminiService:
    def __init__(self, config: AppConfig):
        self.config = config

    def generate_report_data(self, prompt: str) -> dict:
        """
        Asks Gemini for structured JSON content only (no HTML markup at all).
        This dict is later merged into sample_template.html by TemplateService
        using Jinja2 -- Gemini never sees or touches the HTML/CSS, so it can't
        break layout, drop tags, or corrupt styling.
        """
        if not self.config.api_key:
            raise ValueError("Gemini API key is missing. Please add it to your Cred.env file.")

        model = genai.GenerativeModel(
            'gemini-2.5-flash',
            generation_config={"response_mime_type": "application/json"},
        )
        response = model.generate_content(prompt)

        raw = response.text.strip()
        # response_mime_type=application/json should return clean JSON, but
        # strip markdown fences defensively in case the model still adds them.
        if raw.startswith("```"):
            raw = raw.split("```", 2)[1]
            if raw.startswith("json"):
                raw = raw[4:]
        raw = raw.strip().rstrip("`").strip()

        try:
            return json.loads(raw)
        except json.JSONDecodeError as e:
            raise ValueError(f"Gemini did not return valid JSON: {e}\n--- raw response ---\n{raw[:2000]}")


class AnthropicService:
    def __init__(self, config: AppConfig):
        self.config = config

    def generate_report_data(self, prompt: str) -> dict:
        """
        Asks Anthropic Claude for structured JSON content only (no HTML markup).
        This dict is later merged into sample_template.html by TemplateService
        using Jinja2 -- Claude never sees or touches the HTML/CSS.
        """
        if not self.config.anthropic_api_key:
            raise ValueError("Anthropic API key is missing. Please add it to your Cred.env file.")

        import anthropic
        client = anthropic.Anthropic(api_key=self.config.anthropic_api_key)
        model_name = self.config.claude_model or "claude-sonnet-4-6"

        system_prompt = (
            "You are an expert Australian real estate research analyst. "
            "Generate the requested suburb report content strictly adhering to the JSON schema provided in the user prompt. "
            "You MUST respond ONLY with a single valid JSON object. Do NOT include markdown code blocks, conversational text, or HTML markup."
        )

        response = client.messages.create(
            model=model_name,
            max_tokens=8192,
            system=system_prompt,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = response.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```", 2)[1]
            if raw.startswith("json"):
                raw = raw[4:]
        raw = raw.strip().rstrip("`").strip()

        # Find outer-most JSON object if any leading/trailing text exists
        first_brace = raw.find("{")
        last_brace = raw.rfind("}")
        if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
            raw = raw[first_brace:last_brace+1]

        try:
            return json.loads(raw)
        except json.JSONDecodeError as e:
            raise ValueError(f"Claude did not return valid JSON: {e}\n--- raw response ---\n{raw[:2000]}")


class HtagService:
    ENDPOINT = "https://agent.htagai.com/micro-agents/agents/suburb-analysis/execute"
    _cache = {}

    def __init__(self, config: AppConfig):
        self.config = config

    def fetch_suburb_analysis(self, suburb: str, state: str = "", postcode: str = "", property_type: str = "house", force_refresh: bool = False) -> dict:
        """
        Calls HTAG micro-agent API endpoint to retrieve rich real-time suburb intelligence.
        Caches results in memory to optimize latency and token credits.
        """
        if not self.config.htag_api_key:
            raise ValueError("HTAG API key is missing. Please add it to your Cred.env file.")

        import requests

        # Normalize property type to allowed enum: 'house', 'unit', 'townhouse', 'land'
        pt_normalized = "house"
        if property_type:
            pt_clean = property_type.lower().strip()
            if "unit" in pt_clean or "apartment" in pt_clean or "flat" in pt_clean:
                pt_normalized = "unit"
            elif "townhouse" in pt_clean or "semi" in pt_clean or "terrace" in pt_clean:
                pt_normalized = "townhouse"
            elif "land" in pt_clean:
                pt_normalized = "land"
            else:
                pt_normalized = "house"

        cache_key = f"{suburb.strip().lower()}_{state.strip().lower()}_{str(postcode).strip()}_{pt_normalized}"
        if not force_refresh and cache_key in self._cache:
            return self._cache[cache_key]

        payload = {
            "suburb": suburb.strip(),
            "property_type": pt_normalized
        }
        if state:
            payload["state"] = state.strip().upper()
        if postcode:
            payload["postcode"] = str(postcode).strip()

        headers = {
            "x-api-key": self.config.htag_api_key,
            "Authorization": f"Bearer {self.config.htag_api_key}",
            "Content-Type": "application/json"
        }

        try:
            response = requests.post(self.ENDPOINT, headers=headers, json=payload, timeout=120)
            if response.status_code != 200:
                error_msg = response.text
                try:
                    err_json = response.json()
                    error_msg = err_json.get("detail", response.text)
                except Exception:
                    pass
                raise RuntimeError(f"HTAG API returned status {response.status_code}: {error_msg}")

            result = response.json()
            self._cache[cache_key] = result
            return result
        except requests.exceptions.Timeout:
            raise TimeoutError("HTAG Suburb Analysis request timed out after 120 seconds. Please try again.")
        except requests.exceptions.RequestException as req_err:
            raise RuntimeError(f"Network communication with HTAG API failed: {req_err}")



class TemplateService:
    """
    Renders sample_template.html (a Jinja2 template) with a data dict.

    Uses the lenient `Undefined` (not `StrictUndefined`): a missing key
    renders as blank instead of raising. This is intentional -- it's what
    lets the data_notice() macro and `{{ x | default(...) }}` fallbacks in
    the template degrade gracefully when Gemini's JSON is missing a field,
    rather than crashing the whole report generation over one gap. Report
    Generation's own validation pass (report_generation.py) is what
    actually catches and flags missing/unavailable sections before they
    ever reach this render step -- this class stays a "dumb" renderer.
    """
    def __init__(self):
        self._env = Environment(loader=BaseLoader(), undefined=Undefined, autoescape=False)

    def render(self, template_source: str, data: dict) -> str:
        template = self._env.from_string(template_source)
        return template.render(**data)


class PdfService:

    # Margin reserved on content pages for the header/footer band above.
    # Must be >= the actual rendered height of the templates, or Chromium
    # clips them. Nudge these if the banner ever looks cramped/cut off.
    CONTENT_MARGIN_TOP = "14mm"
    CONTENT_MARGIN_BOTTOM = "14mm"
    # Left/right stay at 0 — .report-section (in sample_template_fixed.html)
    # already carries its own consistent 12mm inset. Adding page-level
    # margin on top of that would reintroduce the double-margin bug.
    CONTENT_MARGIN_SIDE = "0mm"

    def __init__(self, config: AppConfig):
        self.config = config

    def to_data_uri(self, path: str, mime: str) -> str:
        if not os.path.exists(path):
            return ""
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        return f"data:{mime};base64,{b64}"

    def prepare_html_assets(self, html_code: str) -> str:
        logo_path = self.config.get_asset_path("LOGO.svg")
        house_path = self.config.get_asset_path("House.png")
        
        logo_uri = self.to_data_uri(logo_path, "image/svg+xml")
        house_uri = self.to_data_uri(house_path, "image/png")
        
        if logo_uri:
            html_code = html_code.replace('src="LOGO.svg"', f'src="{logo_uri}"')
        if house_uri:
            html_code = html_code.replace('src="House.png"', f'src="{house_uri}"')
        return html_code

    async def _generate_pdf_playwright_async(self, html_text: str) -> bytes:
        # This expects the report HTML to follow sample_template_fixed.html's
        # structure: one full-bleed title page as `<div class="page">`,
        # followed by all the content sections wrapped in a single
        # `<div class="report-body">`. If the Gemini-filled HTML still uses
        # the old per-section "page page-2" wrapper, this split won't find
        # the right elements — swap the base template used by
        # ReportGenerationComponent (self.session.template_content) over to
        # sample_template_fixed.html first.
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            title_page = await browser.new_page()
            content_page = await browser.new_page()
            try:
                await title_page.set_content(html_text, wait_until="networkidle")
                logo_uri = self.to_data_uri(
                    self.config.get_asset_path("LOGO.svg"), "image/svg+xml"
                )

                HEADER_TEMPLATE = f"""
                <div style="
                    width:100%;
                    height:10mm;
                    margin:0;
                    padding:0;
                    position:absolute;
                    left:0;
                    right:0;
                    bottom:0;
                    box-sizing:border-box;
                    background:#162338;
                    -webkit-print-color-adjust:exact;
                    print-color-adjust:exact;
                ">
                    <table style="
                        width:100%;
                        height:100%;
                        border-collapse:collapse;
                        margin:0;
                        padding:0;
                    ">
                        <tr>
                            <td style="padding-left:12mm; vertical-align:middle;">
                                <img src="{logo_uri}" style="height:8mm;">
                            </td>

                            <td style="
                                padding-right:12mm;
                                text-align:right;
                                vertical-align:middle;
                                color:#ffffff;
                            ">
                                SMARTPROPGUIDE
                            </td>
                        </tr>
                    </table>
                </div>
                """
                FOOTER_TEMPLATE = """
                                <div style="
                                    width:100%;
                                    height:14mm;
                                    margin:0;
                                    padding:0;
                                    position:absolute;
                                    left:0;
                                    right:0;
                                    bottom:0;
                                    box-sizing:border-box;
                                    background:#162338;
                                    color:#ffffff;
                                    font-family:Arial, sans-serif;
                                    font-size:8pt;
                                    -webkit-print-color-adjust:exact;
                                    print-color-adjust:exact;
                                ">
                                    <table style="
                                        width:100%;
                                        height:100%;
                                        border-collapse:collapse;
                                        margin:0;
                                        padding:0;
                                    ">
                                        <tr>
                                            <td style="
                                                padding-left:12mm;
                                                vertical-align:middle;
                                                color:#ffffff;
                                            ">
                                                Provided by SMARTPROPGUIDE
                                            </td>

                                            <td style="
                                                padding-right:12mm;
                                                text-align:right;
                                                vertical-align:middle;
                                                color:#ffffff;
                                            ">
                                                <span class="pageNumber" style="font-size:9pt; font-weight:500; color:#d7b35e; letter-spacing:0.1em;"></span>
                                            </td>
                                        </tr>
                                    </table>
                                </div>
                        <script>
                        
                        window.addEventListener('DOMContentLoaded', () => {
                            const el = document.querySelector('.pageNumber');
                            if (el) { el.textContent = el.textContent.padStart(2, '0'); }
                        });
                        </script>
                        """

                await title_page.add_style_tag(
                    content="""
                    .report-body {
                        display: none !important;
                    }

                    html,
                    body {
                        margin: 0 !important;
                        padding: 0 !important;
                        width: 210mm !important;
                        min-width: 210mm !important;
                        max-width: 210mm !important;
                        background: transparent !important;
                    }

                    .page {
                        display: block !important;
                        width: 210mm !important;
                        height: 297mm !important;
                        min-width: 210mm !important;
                        max-width: 210mm !important;
                        min-height: 297mm !important;
                        max-height: 297mm !important;

                        margin: 0 !important;
                        padding: 0 !important;

                        box-sizing: border-box !important;
                        overflow: hidden !important;
                    }
                    """
                )

                title_pdf_bytes = await title_page.pdf(
                    width="210mm",
                    height="297mm",
                    print_background=True,
                    margin={
                        "top": "0mm",
                        "right": "0mm",
                        "bottom": "0mm",
                        "left": "0mm",
                    },
                    display_header_footer=False,
                )

                

                # ============================================================
                # CONTENT PAGES
                # ============================================================

                await content_page.set_content(html_text, wait_until="networkidle")

                await content_page.add_style_tag(
                    content="""
                    .page {
                        display: none !important;
                    }

                    .report-body {
                        display: block !important;
                    }

                    html,
                    body {
                        margin: 0 !important;
                        padding: 0 !important;
                        width: 210mm !important;
                        min-width: 210mm !important;
                        max-width: 210mm !important;
                    }

                    .report-body {
                        width: 210mm !important;
                        padding-top: 4mm;
                        margin: 0 !important;
                    }

                    .report-section {
                        width: 210mm !important;
                        margin: 0 !important;
                    }

                    .c-section {
                        break-inside: auto !important;
                        page-break-inside: auto !important;
                    }
                    """
                )

                content_pdf_bytes = await content_page.pdf(
                    width="210mm",
                    height="297mm",
                    print_background=True,
                    display_header_footer=True,
                    header_template=HEADER_TEMPLATE,
                    footer_template=FOOTER_TEMPLATE,
                    margin={
                        "top": "14mm",
                        "bottom": "14mm",
                        "left": "0mm",
                        "right": "0mm",
                    },
                )

            finally:
                await title_page.close()
                await content_page.close()

        # --- merge title + content into the final report, in memory ---
        writer = PdfWriter()
        for pdf_bytes in (title_pdf_bytes, content_pdf_bytes):
            reader = PdfReader(BytesIO(pdf_bytes))
            for pg in reader.pages:
                writer.add_page(pg)
        out_buffer = BytesIO()
        writer.write(out_buffer)
        return out_buffer.getvalue()

    def convert_html_to_pdf(self, html_content: str) -> bytes:
        # Prepare HTML with embedded asset data URIs
        html_prepared = self.prepare_html_assets(html_content)
        # Run the async Playwright PDF generation synchronously
        pdf_bytes = asyncio.run(self._generate_pdf_playwright_async(html_prepared))
        return pdf_bytes