# from app import get_pdf_bytes_playwright
import os
import io
import re
import datetime
import pandas as pd
import streamlit as st
from claude_client import generate_html_report
from htag_client import fetch_suburb, AmbiguousSuburb
from abs_client import fetch_census_medians
import asyncio
import base64
from concurrent.futures import ThreadPoolExecutor
from playwright.async_api import async_playwright
from dotenv import load_dotenv
from xhtml2pdf import pisa


# ==============================================================================
# 1. SETUP & CONFIGURATION
# ==============================================================================
st.set_page_config(
    page_title="SmartPropGuid Report Generator",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# abs_path = os.path.dirname(os.path.abspath(__file__))
# logo_path = os.path.join(abs_path, "LOGO.png")
# html_code = html_code.replace('src="LOGO.png"', f'src="{logo_path}"')


# Load environment variables from Cred.env or .env relative to script directory
script_dir = os.path.dirname(os.path.abspath(__file__))
cred_path = os.path.join(script_dir, "Cred.env")
env_path = os.path.join(script_dir, ".env")
print(f"DEBUG: Looking for Cred.env at: {cred_path}")
print(f"DEBUG: File exists: {os.path.exists(cred_path)}")

if os.path.exists(cred_path):
    load_dotenv(cred_path)
else:
    load_dotenv(env_path)


# Initialize Anthropic Claude API
api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
if not api_key:
    st.warning("⚠️ ANTHROPIC_API_KEY not found in Cred.env. Please configure it to enable AI generation.")

htag_key = os.environ.get("HTAG_API_KEY", "").strip()
# Initialize Session State
if "theme" not in st.session_state:
    st.session_state.theme = "dark"

if "generated_report_html" not in st.session_state:
    st.session_state.generated_report_html = None

if "df_data" not in st.session_state:
    st.session_state.df_data = None

if "template_content" not in st.session_state:
    st.session_state.template_content = ""

# Form field defaults
FORM_DEFAULTS = {
    "full_name": "",
    "phone": "",
    "email": "",
    "property_type": "House",
    "suburb": "",
    "budget": "Under $500k",
    "intention": "Live in",
}
for k, v in FORM_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# Toggle Theme Helper
def toggle_theme():
    st.session_state.theme = "light" if st.session_state.theme == "dark" else "dark"

# Helper to start a compact loading box
def start_loader(message: str):
    placeholder = st.empty()
    placeholder.markdown(
        f'''<div class="loader-container">
                <div class="loader"></div>
                <p class="loader-text">{message}</p>
            </div>''',
        unsafe_allow_html=True,
    )
    return placeholder

# Helper to stop the loading box
def stop_loader(placeholder):
    if placeholder:
        placeholder.empty()

IS_DARK = st.session_state.theme == "dark"

# Splits a run of concatenated `<div class="page">...</div>` /
# `<div class="page page-2">...</div>` blocks back into a list of the
# individual page strings. Used to divide the report template into two
# halves for parallel generation, and to pull the pieces back apart again
# once Claude has returned each half.
PAGE_START_RE = re.compile(r'<div class="page(?: page-2)?">')

def split_into_pages(html_fragment):
    starts = [m.start() for m in PAGE_START_RE.finditer(html_fragment)]
    if not starts:
        return [html_fragment]
    starts.append(len(html_fragment))
    return [html_fragment[starts[i]:starts[i + 1]] for i in range(len(starts) - 1)]

# Budget bracket options for the intake form, mapped to their (min, max) bounds
# in dollars (None = open-ended). Single source of truth for both the
# selectbox options and the listings filter below, so the two can never drift.
BUDGET_RANGES = {
    "Under $500k": (None, 500_000),
    "$500k–$800k": (500_000, 800_000),
    "$800k–$1.2M": (800_000, 1_200_000),
    "$1.2M–$1.5M": (1_200_000, 1_500_000),
    "$1.5M–$2M": (1_500_000, 2_000_000),
    "$2M–$2.5M": (2_000_000, 2_500_000),
    "$2.5M–$3M": (2_500_000, 3_000_000),
    "$3M–$4M": (3_000_000, 4_000_000),
    "$4M–$5M": (4_000_000, 5_000_000),
    "Above $5M": (5_000_000, None),
}
BUDGET_OPTIONS = list(BUDGET_RANGES.keys())

# Helper to filter property listing datasets by postcode and budget
def filter_property_data(df, postcode_str, budget_str, property_type_str):
    if df is None or len(df) == 0:
        return df
    df_filtered = df.copy()
    if postcode_str:
        try:
            pc_val = float(postcode_str)
            if 'Property post code' in df_filtered.columns:
                df_filtered = df_filtered[df_filtered['Property post code'] == pc_val]
        except ValueError:
            pass
    if budget_str and budget_str in BUDGET_RANGES and 'Purchase price' in df_filtered.columns:
        low, high = BUDGET_RANGES[budget_str]
        if low is not None:
            df_filtered = df_filtered[df_filtered['Purchase price'] >= low]
        if high is not None:
            df_filtered = df_filtered[df_filtered['Purchase price'] <= high]
    if property_type_str and 'Primary purpose' in df_filtered.columns:
        if property_type_str == "Land":
            df_filtered = df_filtered[df_filtered['Primary purpose'] == 'Vacant land']
        elif property_type_str in ["House", "Unit", "Townhouse"]:
            df_filtered = df_filtered[df_filtered['Primary purpose'] == 'Residence']
    if len(df_filtered) > 50:
        if 'Contract date' in df_filtered.columns:
            try:
                df_filtered = df_filtered.sort_values(by='Contract date', ascending=False)
            except Exception:
                pass
        df_filtered = df_filtered.head(50)
    return df_filtered

# ==============================================================================
# 2. DESIGN SYSTEM & CSS INJECTION
# ==============================================================================
# Color palette definitions depending on the selected theme
BG_COLOR = "#1A2336" if IS_DARK else "#28344D"
BG_SUBTLE = "#0c0c0f" if IS_DARK else "#f9fafb"
CARD_COLOR = "#0c0c0f" if IS_DARK else "#ffffff"
CARD_HOVER = "#131316" if IS_DARK else "#f4f4f5"
BORDER_COLOR = "#1e1e24" if IS_DARK else "#e4e4e7"
BORDER_SUBTLE = "#16161a" if IS_DARK else "#f0f0f2"
TEXT_COLOR = "#fafafa" if IS_DARK else "#09090b"
TEXT_MUTED = "#71717a"
TEXT_DIM = "#52525b" if IS_DARK else "#a1a1aa"
ACCENT_COLOR = "#2563eb"
ACCENT_MUTED = "#1d4ed8"

css = f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    /* Hide Streamlit default components for custom branding */
    header[data-testid="stHeader"], #MainMenu, footer, [data-testid="stToolbar"],
    [data-testid="stDecoration"], [data-testid="stStatusWidget"], .stDeployButton,
    div[data-testid="stSidebarCollapsedControl"] {{
        display: none !important;
    }}

    /* Global App Container */
    html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"], .main, .block-container, section[data-testid="stMain"] {{
        background-color: {BG_COLOR} !important;
        color: {TEXT_COLOR} !important;
        font-family: 'DM Sans', -apple-system, sans-serif !important;
    }}
    .block-container {{
        padding: 2rem 2.5rem 3rem !important;
        max-width: 1300px !important;
    }}

    /* Loader container – small centered box */
    .loader-container {{
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        padding: 1.5rem;
        background: rgba(0,0,0,0.5);
        border-radius: 12px;
        width: 260px;
        margin: auto;
        box-shadow: 0 4px 12px rgba(0,0,0,0.3);
    }}
    .loader {{
        border: 4px solid {BG_SUBTLE};
        border-top: 4px solid {ACCENT_COLOR};
        border-radius: 50%;
        width: 60px;
        height: 60px;
        animation: spin 1s linear infinite;
    }}
    @keyframes spin {{
        0% {{ transform: rotate(0deg); }}
        100% {{ transform: rotate(360deg); }}
    }}
    .loader-text {{
        margin-top: 1rem;
        color: {TEXT_COLOR};
        font-size: 1.1rem;
        text-align: center;
    }}

    /* Tabs (pill-style navigation) */
    button[data-baseweb="tab"] {{
        background: transparent !important;
        color: {TEXT_MUTED} !important;
        font-size: 0.88rem !important;
        font-weight: 500 !important;
        padding: 0.6rem 1.2rem !important;
        border: 1px solid transparent !important;
        border-radius: 8px !important;
        transition: all 0.2s ease !important;
    }}
    button[data-baseweb="tab"]:hover {{
        color: {TEXT_COLOR} !important;
        background: {CARD_HOVER} !important;
    }}
    button[data-baseweb="tab"][aria-selected="true"] {{
        color: {TEXT_COLOR} !important;
        background: {CARD_COLOR} !important;
        border-color: {BORDER_COLOR} !important;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1) !important;
    }}
    [data-baseweb="tab-highlight"], [data-baseweb="tab-border"] {{
        display: none !important;
    }}
    [data-baseweb="tab-list"] {{
        gap: 6px !important;
        background: {BG_SUBTLE} !important;
        border: 1px solid {BORDER_COLOR} !important;
        border-radius: 12px !important;
        padding: 4px;
        margin-bottom: 2rem !important;
    }}

    /* Column spacing */
    [data-testid="stHorizontalBlock"] {{
        gap: 1.5rem !important;
    }}

    /* Bordered card containers — st.container(border=True) */
    div[data-testid="stVerticalBlockBorderWrapper"] {{
        margin-bottom: 1.5rem;
    }}
    div[data-testid="stVerticalBlockBorderWrapper"] > div[data-testid="stVerticalBlock"] {{
        background-color: {CARD_COLOR} !important;
        border: 1px solid {BORDER_COLOR} !important;
        border-radius: 12px !important;
        padding: 1.75rem !important;
        gap: 1rem !important;
        transition: border-color 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }}
    div[data-testid="stVerticalBlockBorderWrapper"] > div[data-testid="stVerticalBlock"]:hover {{
        border-color: {ACCENT_COLOR} !important;
    }}

    /* Brand banner styling */
    .brand {{
        display: flex;
        align-items: center;
        gap: 10px;
        margin-bottom: 1.5rem;
    }}
    .brand-logo {{
        font-size: 1.6rem;
        background: linear-gradient(135deg, #3b82f6, #2563eb);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-weight: 800;
        letter-spacing: -0.05em;
    }}
    .brand-title {{
        font-size: 1.6rem;
        font-weight: 700;
        color: {TEXT_COLOR};
        letter-spacing: -0.03em;
    }}
    .brand-subtitle {{
        font-size: 0.88rem;
        color: {TEXT_MUTED};
        margin-top: -5px;
    }}

    /* Professional details badge */
    .badge {{
        display: inline-block;
        padding: 4px 10px;
        border-radius: 6px;
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.02em;
        text-transform: uppercase;
    }}
    .badge-accent {{
        color: {ACCENT_COLOR};
        background: rgba(37, 99, 235, 0.1);
        border: 1px solid rgba(37, 99, 235, 0.2);
    }}

    /* Streamlit widget overrides — selectbox, text input, multiselect */
    .stTextInput>div>div>input,
    .stTextInput input,
    .stSelectbox>div>div>div,
    .stSelectbox [data-baseweb="select"] > div,
    .stMultiSelect>div>div,
    [data-baseweb="select"] > div {{
        background-color: {BG_SUBTLE} !important;
        border: 1px solid {BORDER_COLOR} !important;
        color: {TEXT_COLOR} !important;
        border-radius: 8px !important;
    }}
    /* Selectbox dropdown option text */
    [data-baseweb="select"] span,
    [data-baseweb="select"] div {{
        color: {TEXT_COLOR} !important;
    }}
    /* Placeholder text color */
    .stTextInput input::placeholder {{
        color: {TEXT_DIM} !important;
        opacity: 1 !important;
    }}
    
    /* Styled HTML Tables */
    .data-table {{
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        font-size: 0.85rem;
        margin-top: 1rem;
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid {BORDER_COLOR};
    }}
    .data-table th {{
        background: {BG_SUBTLE};
        color: {TEXT_MUTED};
        text-align: left;
        padding: 0.75rem 1rem;
        font-size: 0.78rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        border-bottom: 1px solid {BORDER_COLOR};
    }}
    .data-table td {{
        padding: 0.8rem 1rem;
        color: {TEXT_COLOR};
        background: {CARD_COLOR};
        border-bottom: 1px solid {BORDER_SUBTLE};
    }}
    .data-table tr:last-child td {{
        border-bottom: none;
    }}
    
    /* Previews */
    .preview-box {{
        background-color: {BG_SUBTLE};
        border: 1px solid {BORDER_COLOR};
        border-radius: 8px;
        padding: 1rem;
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.8rem;
        overflow-x: auto;
        white-space: pre-wrap;
        color: {TEXT_COLOR};
    }}

    /* Theme toggle button — white bg + black text in light, dark bg + white text in dark */
    [data-testid="stBaseButton-secondary"] button,
    button[kind="secondary"],
    .stButton > button {{
        background-color: {"#1e1e24" if IS_DARK else "#ffffff"} !important;
        color: {TEXT_COLOR} !important;
        border: 1px solid {BORDER_COLOR} !important;
    }}
    [data-testid="stBaseButton-secondary"] button *,
    [data-testid="stBaseButton-secondary"] button p,
    [data-testid="stBaseButton-secondary"] button span,
    .stButton > button p,
    .stButton > button span {{
        color: {TEXT_COLOR} !important;
    }}

    /* Widget labels (selectbox, text_input, checkbox) — readable in both modes */
    .stSelectbox label,
    .stTextInput label,
    .stCheckbox label,
    .stCheckbox label p,
    .stCheckbox span,
    .stCheckbox p,
    [data-testid="stCheckbox"] label,
    [data-testid="stCheckbox"] label p,
    [data-testid="stCheckbox"] span {{
        color: {TEXT_COLOR} !important;
    }}

    /* Checkbox box (the square) — border and background match theme */
    [data-baseweb="checkbox"] > div:first-child,
    [data-testid="stCheckbox"] [data-baseweb="checkbox"] > div {{
        background-color: {BG_SUBTLE} !important;
        border-color: {BORDER_COLOR} !important;
    }}
    /* Checked state — keep accent blue */
    [data-baseweb="checkbox"][aria-checked="true"] > div:first-child {{
        background-color: {ACCENT_COLOR} !important;
        border-color: {ACCENT_COLOR} !important;
    }}
</style>
"""
st.markdown(css, unsafe_allow_html=True)

# ==============================================================================
# 3. HEADER & BRAND AREA
# ==============================================================================
# Load the logo mark for the header (falls back to a text glyph if missing)
logo_path = os.path.join(script_dir, "LOGO.png")
if os.path.exists(logo_path):
    with open(logo_path, "rb") as f:
        logo_html = f'<img src="data:image/png;base64,{base64.b64encode(f.read()).decode("utf-8")}" style="height:44px; width:44px; border-radius:10px; object-fit:cover;" alt="SmartPropGuid logo" />'
else:
    logo_html = '<span class="brand-logo">◆</span>'

head_left, head_right = st.columns([9, 2])
with head_left:
    st.markdown(f"""
    <div class="brand">
        {logo_html}
        <div>
            <span class="brand-title">SmartPropGuid Report Engine</span>
            <div class="brand-subtitle">Pre-Sales Manual Compilation & AI Generation Tool</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
with head_right:
    theme_label = "☀️ Light Mode" if IS_DARK else "🌙 Dark Mode"
    st.button(theme_label, on_click=toggle_theme, use_container_width=True)

st.markdown("<hr style='margin-top:0.5rem; margin-bottom:1.5rem; border-color:" + BORDER_COLOR + "'>", unsafe_allow_html=True)

# ==============================================================================
# 4. APPLICATION TABS (FLOW STEPS)
# ==============================================================================
tab_preferences, tab_upload, tab_generate = st.tabs([
    "📋 1. Customer Preferences", 
    "📂 2. Data & Template Upload", 
    "✨ 3. AI Report Generation"
])



# ------------------------------------------------------------------------------
# TAB 1: CUSTOMER PREFERENCES
# ------------------------------------------------------------------------------
with tab_preferences:
    st.markdown("### Customer Preferences Form")
    st.markdown("Capture customer search requirements to guide the AI report writer.")
    
    # --- Customer Info Card ---
    with st.container(border=True):
        st.markdown("<h4>Customer Information</h4>", unsafe_allow_html=True)
        ci_col1, ci_col2, ci_col3 = st.columns(3)
        with ci_col1:
            st.text_input("Full Name", placeholder="e.g. John Smith", key="full_name")
        with ci_col2:
            st.text_input("Phone Number", placeholder="e.g. 0412 345 678", key="phone")
        with ci_col3:
            st.text_input("Email Address", placeholder="e.g. john@email.com", key="email")

    col1, col2 = st.columns(2)

    with col1:
        with st.container(border=True):
            st.markdown("<h4>Property Details</h4>", unsafe_allow_html=True)

            property_type = st.selectbox(
                "What type of property are you looking for?",
                options=["House", "Unit", "Townhouse", "Land", "Not sure"],
                key="property_type"
            )

            suburb = st.text_input(
                "Which suburb or area are you interested in?",
                placeholder="e.g. Richmond, VIC 3121 or 2000",
                key="suburb"
            )

            budget = st.selectbox(
                "What is your budget?",
                options=BUDGET_OPTIONS,
                key="budget"
            )

            intention = st.selectbox(
                "Are you buying to live in or invest?",
                options=["Live in", "Invest", "Both"],
                key="intention"
            )

    with col2:
        with st.container(border=True):
            st.markdown("<h4>Sub-regional Priorities & Preferences</h4>", unsafe_allow_html=True)

            priorities_list = [
                "Good schools nearby",
                "Public transport access",
                "Shopping centres nearby",
                "Parks and green spaces",
                "Hospital or medical centre nearby",
                "Low flood risk",
                "Low bushfire risk",
                "Quiet neighbourhood",
                "Investment potential",
                "Family friendly area",
                "Close to CBD"
            ]

            st.markdown("<p style='font-size:0.85rem; color:"+TEXT_MUTED+"; margin-bottom:10px;'>Tap to select the features that matter most:</p>", unsafe_allow_html=True)

            # Toggle-chip multi-select (replaces the old 2-column checkbox grid)
            selected_priorities = st.pills(
                "Priorities",
                options=priorities_list,
                selection_mode="multi",
                default=[],
                label_visibility="collapsed",
                key="priorities_pills",
            ) or []

    # --- Reset & Submit Buttons ---
    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
    btn_col1, btn_col2, btn_spacer = st.columns([1, 1, 4])
    with btn_col1:
        if st.button("🔄 Reset", use_container_width=True, key="reset_btn"):
            for k in FORM_DEFAULTS.keys():
                if k in st.session_state:
                    del st.session_state[k]
            if "priorities_pills" in st.session_state:
                del st.session_state["priorities_pills"]
            st.rerun()
    with btn_col2:
        if st.button("✅ Submit", use_container_width=True, key="submit_btn"):
            full_name = st.session_state.get("full_name", "").strip()
            phone = st.session_state.get("phone", "").strip()
            email = st.session_state.get("email", "").strip()
            suburb_input = st.session_state.get("suburb", "").strip()
            
            if not full_name or not phone or not email or not suburb_input:
                st.error("❌ Full Name, Phone Number, Email Address, and Suburb/Area are required fields.")
            else:
                EXCEL_PATH = r"C:\Users\ahmma\.gemini\antigravity-ide\scratch\SmartPropGuid_Report_Gen_Gemini\SPG_Customer_Intake_Form.xlsx"
                try:
                    import openpyxl
                    import re
                    from copy import copy
                    
                    if not os.path.exists(EXCEL_PATH):
                        st.error(f"❌ Excel file not found: {EXCEL_PATH}. Please make sure the template file exists.")
                    else:
                        wb = openpyxl.load_workbook(EXCEL_PATH)
                        sheet = wb.active
                        
                        # Find the last row with data in Column A (Submission ID)
                        last_row = 3
                        for r in range(4, sheet.max_row + 2):
                            if sheet.cell(row=r, column=1).value is not None:
                                last_row = r
                        
                        # Generate Submission ID
                        last_id = sheet.cell(row=last_row, column=1).value
                        if last_id and isinstance(last_id, str) and last_id.startswith("SPG-"):
                            try:
                                num = int(last_id.split("-")[1])
                                next_id = f"SPG-{num + 1:03d}"
                            except (IndexError, ValueError):
                                next_id = "SPG-002"
                        else:
                            next_id = "SPG-001"
                        
                        # Suburb, Postcode, and State parsing
                        suburb_clean, postcode_clean, state_clean = "", "", ""
                        if suburb_input:
                            postcode_match = re.search(r"\b\d{3,4}\b", suburb_input)
                            postcode_clean = postcode_match.group(0) if postcode_match else ""
                            
                            state_match = re.search(r"\b(VIC|NSW|QLD|WA|SA|TAS|ACT|NT)\b", suburb_input, re.IGNORECASE)
                            state_clean = state_match.group(0).upper() if state_match else ""
                            
                            suburb_clean = suburb_input
                            if postcode_clean:
                                suburb_clean = suburb_clean.replace(postcode_clean, "")
                            if state_clean:
                                suburb_clean = re.sub(rf"\b{state_clean}\b", "", suburb_clean, flags=re.IGNORECASE)
                                
                            suburb_clean = re.sub(r"[,\-\s]+", " ", suburb_clean).strip()
                        
                        # Map priority chip selections to Yes/No
                        priorities_yes_no = ["Yes" if p in selected_priorities else "No" for p in priorities_list]
                        
                        next_row = last_row + 1
                        
                        # Date format
                        date_submitted = datetime.date.today().strftime("%d/%m/%Y")
                        
                        # Assemble row data (now 25 columns)
                        row_values = [
                            next_id,                           # 1: Submission ID
                            date_submitted,                    # 2: Date Submitted
                            full_name,                         # 3: Full Name
                            email,                             # 4: Email Address
                            phone,                             # 5: Phone Number
                            st.session_state.get("property_type", "House"),  # 6: Property Type
                            suburb_clean,                      # 7: Suburb / Area
                            postcode_clean,                    # 8: Postcode
                            state_clean,                       # 9: State
                            st.session_state.get("budget", ""), # 10: Budget Range
                            st.session_state.get("intention", ""), # 11: Buying Purpose
                        ]
                        # Append 11 priorities
                        row_values.extend(priorities_yes_no)
                        # Append remaining columns: Additional Notes (23), Report Status (24), Assigned To (25)
                        row_values.extend(["", "Pending", "Shoyeb"])
                        
                        # Write to the cells and copy style if last_row has styles
                        for col_idx, val in enumerate(row_values, start=1):
                            new_cell = sheet.cell(row=next_row, column=col_idx, value=val)
                            if last_row >= 4:
                                src_cell = sheet.cell(row=last_row, column=col_idx)
                                if src_cell.has_style:
                                    new_cell.font = copy(src_cell.font)
                                    new_cell.border = copy(src_cell.border)
                                    new_cell.fill = copy(src_cell.fill)
                                    new_cell.number_format = copy(src_cell.number_format)
                                    new_cell.protection = copy(src_cell.protection)
                                    new_cell.alignment = copy(src_cell.alignment)
                        
                        wb.save(EXCEL_PATH)
                        st.success(f"✅ Submission saved to Excel successfully as {next_id}!")
                except Exception as e:
                    st.error(f"❌ Failed to save: {e}")

# ------------------------------------------------------------------------------
# TAB 2: DATA & TEMPLATE UPLOAD
# ------------------------------------------------------------------------------
with tab_upload:
    st.markdown("### Manual Data Compilation & Predefined Layouts")
    st.markdown("Upload the source documents gathered by your operator for Step 3.")
    
    col_data, col_template = st.columns(2)
    
    with col_data:
        with st.container(border=True):
            st.markdown("<h4>1. Source Data File (CSV / Excel)</h4>", unsafe_allow_html=True)
            st.markdown(f"<p style='font-size:0.8rem; color:{TEXT_MUTED};'>Upload the property listings, demographics or market statistics file compiled manually.</p>", unsafe_allow_html=True)

            uploaded_data = st.file_uploader(
                "Select compiled data file",
                type=["csv", "xlsx", "xls"],
                key="data_uploader"
            )

            # Read and display data preview
            data_preview_html = ""
            if uploaded_data is not None:
                try:
                    uploaded_data.seek(0)
                    if uploaded_data.name.endswith(".csv"):
                        st.session_state.df_data = pd.read_csv(uploaded_data)
                    else:
                        st.session_state.df_data = pd.read_excel(uploaded_data)

                    st.success(f"Successfully loaded: `{uploaded_data.name}` ({len(st.session_state.df_data)} rows)")
                    st.markdown("##### File Preview (First 5 Rows):")
                    st.dataframe(st.session_state.df_data.head(5), use_container_width=True)

                    # Format to a table string for the LLM
                    data_preview_html = st.session_state.df_data.head(20).to_html(index=False, classes="data-table")
                except Exception as e:
                    st.error(f"Error reading file: {e}")
            else:
                st.session_state.df_data = None

            df_data = st.session_state.df_data

    with col_template:
        with st.container(border=True):
            st.markdown("<h4>2. Predefined Report Template (HTML)</h4>", unsafe_allow_html=True)
            st.markdown(f"<p style='font-size:0.8rem; color:{TEXT_MUTED};'>Upload your customized corporate report HTML template. Leave empty to use the system default template.</p>", unsafe_allow_html=True)

            uploaded_template = st.file_uploader(
                "Select custom HTML template",
                type=["html", "htm"],
                key="template_uploader"
            )

            # Load custom or default template
            default_template_path = os.path.join(script_dir, "sample_template.html")

            if uploaded_template is not None:
                try:
                    uploaded_template.seek(0)
                    st.session_state.template_content = uploaded_template.read().decode("utf-8")
                    st.success(f"Custom template loaded: `{uploaded_template.name}`")
                except Exception as e:
                    st.error(f"Error reading template: {e}")
            else:
                if os.path.exists(default_template_path):
                    with open(default_template_path, "r", encoding="utf-8") as f:
                        st.session_state.template_content = f.read()
                    st.info("ℹ️ Using default SmartPropGuid template.")
                else:
                    st.session_state.template_content = ""
                    st.warning("⚠️ System default template not found. Please upload a template.")

            template_content = st.session_state.template_content

            # Template preview in a collapsible expander
            if template_content:
                with st.expander("👁️ View Template HTML Structure"):
                    st.markdown(f'<div class="preview-box">{template_content}</div>', unsafe_allow_html=True)

    # Custom AI System Prompt
    with st.container(border=True):
        st.markdown("<h4>3. Pre-Sales Operator Instructions (Prompt)</h4>", unsafe_allow_html=True)
        custom_prompt = st.text_area(
            "Define what aspects you want the AI to emphasize in the report analysis",
            value="Focus heavily on capital growth trends, school catchment boundaries, and transport proximity recommendations based on the customer requirements and listing data.",
            height=100
        )

# ------------------------------------------------------------------------------
# TAB 3: AI GENERATION & PDF EXPORT
# ------------------------------------------------------------------------------
with tab_generate:
    st.markdown("### Generate and Review Report")
    
    # Check if API key is present
    if not api_key:
        st.error("❌ Cannot generate report: Anthropic API key is missing. Please add ANTHROPIC_API_KEY to your Cred.env file.")
    else:
        # Layout: Settings Summary Card & Generation Button
        sum_col1, sum_col2 = st.columns([3, 1])
        with sum_col1:
            st.markdown(f"""
            <div style="background-color: {BG_SUBTLE}; padding: 1rem; border-radius: 8px; border: 1px solid {BORDER_COLOR}; font-size: 0.88rem;">
                <strong>Target Area:</strong> {suburb if suburb else "Not specified"} | 
                <strong>Property Type:</strong> {property_type} | 
                <strong>Budget:</strong> {budget} | 
                <strong>Purpose:</strong> {intention} | 
                <strong>Key Preferences Selected:</strong> {', '.join(selected_priorities) if selected_priorities else "None"}
            </div>
            """, unsafe_allow_html=True)
        with sum_col2:
            generate_btn = st.button("✨ Generate AI Report", type="primary", use_container_width=True)

        if generate_btn:
            # Get postcode from suburb input
            suburb_input = st.session_state.get("suburb", "").strip()
            postcode_str = ""
            if suburb_input:
                import re
                pm = re.search(r"\b\d{3,4}\b", suburb_input)
                postcode_str = pm.group(0) if pm else ""

            # Try to auto-load from Property_Data_Split if no manual file is uploaded
            df_active = st.session_state.get("df_data")
            if df_active is None and postcode_str:
                try:
                    pc_val = float(postcode_str)
                    split_dir = r"C:\Users\ahmma\Desktop\Property_Data_Split"
                    if os.path.exists(split_dir):
                        for fname in os.listdir(split_dir):
                            if fname.startswith("postcode_") and fname.endswith(".csv"):
                                parts = fname.replace("postcode_", "").replace(".csv", "").split("_to_")
                                if len(parts) == 2:
                                    start_pc = float(parts[0])
                                    end_pc = float(parts[1])
                                    if start_pc <= pc_val <= end_pc:
                                        csv_path = os.path.join(split_dir, fname)
                                        df_active = pd.read_csv(csv_path)
                                        st.info(f"ℹ️ Automatically loaded postcode dataset: `{fname}`")
                                        break
                except Exception as e:
                    st.warning(f"⚠️ Could not auto-load postcode dataset: {e}")

            # Filter data to make prompt compact and relevant
            df_filtered = None
            if df_active is not None:
                df_filtered = filter_property_data(
                    df_active, 
                    postcode_str, 
                    st.session_state.get("budget", ""), 
                    st.session_state.get("property_type", "")
                )

            # Prepare instructions and data details to inject into Gemini prompt
            data_context = ""
            source_data_rows = ""
            if df_filtered is not None and len(df_filtered) > 0:
                data_context = f"Manual Listings Data Compiled (Filtered for Postcode {postcode_str}):\n{df_filtered.to_string(index=False)}"
                # Build HTML rows for the report
                rows = []
                for _, r in df_filtered.iterrows():
                    address = r.get('Address') or r.get('Property address') or ''
                    price = r.get('Purchase price') or r.get('Price') or ''
                    rows.append(f"<tr><td>{address}</td><td>{price}</td><td><span class='badge badge-blue'>Score</span></td><td><span class='badge badge-green'>Available</span></td></tr>")
                source_data_rows = "".join(rows)
            else:
                st.warning("No matching listings data found. Report will be generated without property rows.")
                data_context = "No listing data uploaded or matching the criteria."

            # Inject rows into the selected HTML template
            filled_template = st.session_state.template_content.replace("{{ source_data_rows }}", source_data_rows)

            # Split the template into its static <head> (fonts + the large inline
            # <style> block) and the <body> content that actually needs per-suburb
            # data. Only the body is sent to the model — the report template is
            # ~90K characters and asking Claude to also reproduce the CSS verbatim
            # on every request risks the response getting cut off by the token
            # limit before any real page content is written (a truncated response
            # mid-<style> is exactly what renders as a blank white PDF page).
            body_start_tag = "<body>"
            body_end_tag = "</body>"
            if body_start_tag in filled_template and body_end_tag in filled_template:
                head_html = filled_template[: filled_template.index(body_start_tag) + len(body_start_tag)]
                body_inner = filled_template[
                    filled_template.index(body_start_tag) + len(body_start_tag) : filled_template.index(body_end_tag)
                ]
                tail_html = filled_template[filled_template.index(body_end_tag):]
            else:
                # Custom uploaded template with no <body> tag — fall back to sending it whole.
                head_html, body_inner, tail_html = "", filled_template, ""

            # Build Prompt — shared by both chunks below
            def build_prompt(page_html, page_names):
                return f"""
                You are a professional property investment analyst assistant. The report
                you are writing is read directly by the client (a home buyer) — not by
                another analyst.
                You must populate the HTML report pages provided below based on the
                following Inputs. This is one part of a larger report; the pages included
                here, in this exact order, are: {page_names}.

                --- INPUTS ---
                1. Property Type Preference: {property_type}
                2. Target Suburb/Area: {suburb}
                3. Budget: {budget}
                4. Purchase Intention: {intention}
                5. Key Client Priorities: {', '.join(selected_priorities) if selected_priorities else "General property advice"}
                6. Pre-Sales Operator Instructions: {custom_prompt}

                7. Source Data:
                {data_context}

                --- HTML PAGES: {page_names} ---
                {page_html}

                --- WRITING STYLE (this is client-facing, not an investor memo) ---
                a. Plain, warm, everyday English — no investor jargon (e.g. avoid "yield
                   compression", "capital velocity"), no unexplained acronyms.
                b. Never leave a number, score, or recommendation to speak for itself —
                   follow it with a short, plain-English reason it matters to THIS buyer.
                   For example, don't just write "72% auction clearance"; write "72%
                   auction clearance — meaning sellers currently have the upper hand, so
                   be ready to move quickly on a property you like."
                c. If a figure isn't backed by the Source Data, present it as a clearly
                   reasonable estimate ("typically around...") rather than false
                   precision — but always give one, so the report never reads as empty.

                --- COMPILING INSTRUCTIONS ---
                1. If the cover page (`<div class="page">`) is included above, populate its
                   placeholders exactly: `{{ suburb }}`, `{{ postcode }}`, `{{ median_price }}`,
                   `{{ clearance_rate }}`, `{{ dom }}`.
                2. Every `<div class="page page-2">` page is a fully worked EXAMPLE for a
                   fictional suburb ("Woolloomooloo") — it exists only to show structure,
                   section order, tone, and level of numeric detail. Replace EVERY suburb
                   name, statistic, percentage, chart data point/coordinate, list item, and
                   sentence with real content for the ACTUAL target suburb and inputs above.
                   Do not leave any "Woolloomooloo" facts or numbers in the output.
                3. Where exact data isn't available, produce reasonable, internally
                   consistent estimates (a bar's width % must match the number shown next
                   to it; an SVG chart's polyline/circle coordinates must match its axis
                   labels and endpoint text).
                4. The "What's Nearby" amenity counts (cafés, supermarkets, parks, gyms,
                   hospitals, transport stops) are never a single exact number — always
                   give a plausible RANGE instead (e.g. "4–6", "38–45"), since these are
                   estimates, not a verified count.
                5. Keep the HTML tag structure, class names, and inline SVG structure
                   exactly as given — only change text, numbers, and coordinate/width
                   attributes. Do not add, remove, or reorder `<div class="page ...">` sections.
                6. Return ONLY the page div(s) shown above, in the same order, rewritten
                   with real data. Do NOT include `<!doctype html>`, `<html>`, `<head>`,
                   `<style>`, `<body>` or `</body>`/`</html>` tags.
                7. Do NOT wrap the code in markdown blocks like ````html```` or add any
                   conversational intro/outro text. Output only raw HTML.
                """

            # Split the 8 pages into two independently-writable halves and generate
            # them concurrently — this is the main lever on wall-clock time, since a
            # single call previously had to both read and re-write the full ~13,000
            # tokens of report HTML in one uninterrupted stream. Chunk A keeps every
            # page that repeats a headline figure (median price, clearance rate,
            # growth %, price history) in ONE call so those numbers stay internally
            # consistent; Chunk B is the self-contained, non-financial half (lifestyle,
            # community, schools, risk) and carries no dependency on Chunk A's numbers.
            pages = split_into_pages(body_inner)
            if len(pages) == 8:
                chunk_a_pages = [pages[0], pages[1], pages[2], pages[7]]
                chunk_b_pages = [pages[3], pages[4], pages[5], pages[6]]
            else:
                # Custom uploaded template with a different page count/shape —
                # fall back to a single, unsplit call.
                chunk_a_pages, chunk_b_pages = pages, []

            chunk_a_prompt = build_prompt(
                "".join(chunk_a_pages), "Cover, Suburb Snapshot, Growth Outlook, The Verdict"
            )
            chunk_b_prompt = (
                build_prompt(
                    "".join(chunk_b_pages),
                    "Lifestyle & Amenities, Community Profile, Schools & Family Fit, Risk & Planning",
                )
                if chunk_b_pages
                else None
            )

            loader_placeholder = start_loader(
                "Claude is analyzing data and writing the report (two sections in parallel)…"
                if chunk_b_prompt
                else "Claude is analyzing data and populating report layout…"
            )
            try:
                if chunk_b_prompt:
                    with ThreadPoolExecutor(max_workers=2) as executor:
                        future_a = executor.submit(generate_html_report, chunk_a_prompt)
                        future_b = executor.submit(generate_html_report, chunk_b_prompt)
                        generated_a = future_a.result()
                        generated_b = future_b.result()

                    # Chunk A comes back as [Cover, Snapshot, Growth, Verdict] — pull the
                    # Verdict page back out so the final order is pages 1..8.
                    a_split = split_into_pages(generated_a)
                    if len(a_split) == 4:
                        generated_body = "".join(a_split[:3]) + generated_b + a_split[3]
                    else:
                        # Model didn't preserve the exact page boundaries — fall back to
                        # a best-effort concatenation rather than failing the report.
                        generated_body = generated_a + generated_b
                else:
                    generated_body = generate_html_report(chunk_a_prompt)

                # Each chunk only sees its own 3-4 pages, so a model asked to number
                # a footer ("01", "02", ...) numbers it relative to what it was shown,
                # not the final 8-page document — e.g. the Verdict page (last overall)
                # can come back labelled "03" because it was the 3rd footer-bearing
                # page inside chunk A. Renumber deterministically in final page order
                # instead of trusting either chunk's own count.
                footer_counter = {"n": 0}
                def _renumber_footer(match):
                    footer_counter["n"] += 1
                    return f'<span class="p2-footer-page">{footer_counter["n"]:02d}</span>'
                generated_body = re.sub(r'<span class="p2-footer-page">\d+</span>', _renumber_footer, generated_body)

                # Stitch the (model-written) body back into the untouched, verbatim
                # head/CSS and tail so the design can never be corrupted or truncated away.
                clean_html = f"{head_html}\n{generated_body}\n{tail_html}" if head_html else generated_body
                st.session_state.generated_report_html = clean_html
                st.success(" ✅ Report Generated Successfully")

            except Exception as e:
                st.error(f"Failed to generate report from Claude API: {e}")
            finally:
                stop_loader(loader_placeholder)

        # Render generated report and PDF download option if exists
        if st.session_state.generated_report_html:
            st.markdown("### Generated Report Preview")
            
            # PDF Generation Block
            html_code = st.session_state.generated_report_html
            
            # Compiling helper to convert HTML to PDF bytes
            # def get_pdf_bytes(html_text):
            #     pdf_io = io.BytesIO()
            #     pisa_status = pisa.CreatePDF(html_text, dest=pdf_io)
            #     if pisa_status.err:
            #         return None
            #     return pdf_io.getvalue()

            def to_data_uri(path, mime):
                with open(path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode("utf-8")
                return f"data:{mime};base64,{b64}"

            logo_uri = to_data_uri(os.path.join(script_dir, "LOGO.png"), "image/png")
            house_uri = to_data_uri(os.path.join(script_dir, "House.png"), "image/png")

            html_code = html_code.replace('src="LOGO.png"', f'src="{logo_uri}"')
            html_code = html_code.replace('src="House.png"', f'src="{house_uri}"')

            async def get_pdf_bytes_playwright(html_text):
                async with async_playwright() as p:
                    browser = await p.chromium.launch()
                    try:
                        page = await browser.new_page()
                        await page.set_content(html_text, wait_until="networkidle")
                        
                        pdf_bytes = await page.pdf(
                            format="A4",
                            print_background=True,
                            margin={"top": "0px", "right": "0px", "bottom": "14mm", "left": "0px"},
                            display_header_footer=True,
                            header_template="<span></span>",  # empty — suppresses Chromium's default title/url header
                            footer_template="""
                                <div style="width:100%; font-family:'Helvetica Neue', Helvetica, Arial, sans-serif;
                                            font-size:8pt; color:#71717a; text-align:right; padding-right:12mm;">
                                    Page <span class="pageNumber"></span> of <span class="totalPages"></span>
                                </div>
                            """,
                        )
                        return pdf_bytes
                    finally:
                        # This ensures the browser ALWAYS closes, even if an error occurs
                        await browser.close()

            # In Streamlit, use asyncio to run this:
            # pdf_bytes = asyncio.run(get_pdf_bytes_playwright(html_code))
                
            
            pdf_bytes = asyncio.run(get_pdf_bytes_playwright(html_code))


            
            # Download Button
            if pdf_bytes:
                st.download_button(
                    label="📥 Download PDF Report",
                    data=pdf_bytes,
                    file_name=f"SmartPropGuid_Report_{suburb.replace(' ', '_') if suburb else 'General'}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            else:
                st.error("Could not compile HTML to PDF. Check if the HTML template format has errors.")

            # On-screen preview using HTML component iframe
            st.components.v1.html(html_code, height=700, scrolling=True)
